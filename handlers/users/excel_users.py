from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

# getting data and creating virtual xlsx file for download
from utils.db_api.commands import check_user_register


async def export_users_registered(data, course_id):
    wb = Workbook()

    # # get active sheet
    ws = wb.active
    ws.title = "Users List"

    # extend columns size
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    # creating list for excel file column headers
    headings = ["Ism Familiya", "Telefon raqam", "Manzil", "Ro'xatdan o'tgan vaqti"]
    ws.append(headings)

    # changing column header text to bold
    for col in range(1, 10):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)

    i = 2
    for user in data:

        user_register_time = await check_user_register(int(course_id), int(user["telegram_id"]))
        # adding order items to the workbook

        # getting current sheet and putting the text to center
        current = ws[f"A{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        # adding value
        ws[f"A{i}"].value = user["full_name"]

        current = ws[f"B{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"B{i}"].value = user["phone_number"]

        current = ws[f"C{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"C{i}"].value = user["location"]

        current = ws[f"D{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"D{i}"].value = str(user_register_time["created_at"])[0:19]

        i += 1
    else:
        pass

    # creating virtual file for download
    # by this server can send file without saving file
    workbook = save_virtual_workbook(wb)
    # encode it

    return workbook


async def export_users_registered_bot(data):
    wb = Workbook()

    # # get active sheet
    ws = wb.active
    ws.title = "Users List"

    # extend columns size
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # creating list for excel file column headers
    headings = ["Ism Familiya", "Telefon raqam", "Manzil", "Ro'xatdan o'tgan vaqti"]
    ws.append(headings)

    # changing column header text to bold
    for col in range(1, 10):
        ws[get_column_letter(col) + '1'].font = Font(bold=True)

    i = 2
    for user in data:
        # adding order items to the workbook

        # getting current sheet and putting the text to center
        current = ws[f"A{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        # adding value
        ws[f"A{i}"].value = user["full_name"]

        current = ws[f"B{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"B{i}"].value = user["phone_number"]

        current = ws[f"C{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"C{i}"].value = user["location"]

        current = ws[f"D{i}"]
        current.alignment = Alignment(horizontal='center', vertical='center')
        ws[f"D{i}"].value = str(user["created_at"])[0:19]

        i += 1
    else:
        pass

    # creating virtual file for download
    # by this server can send file without saving file
    workbook = save_virtual_workbook(wb)
    # encode it

    return workbook
